import xlsxwriter
import openpyxl

from sale_out.database import CBase_2021_quadra_workout


def demo_writer():# Create an new Excel file and add a worksheet.
    workbook = xlsxwriter.Workbook('demo.xlsx')
    worksheet = workbook.add_worksheet()

    # Widen the first column to make the text clearer.
    worksheet.set_column('A:A', 20)

    # Add a bold format to use to highlight cells.
    bold = workbook.add_format({'bold': True})

    # Write some simple text.
    worksheet.write('A1', 'Hello')

    # Text with formatting.
    worksheet.write('B1', 'World', bold)

    # Write some numbers, with row/column notation.
    worksheet.write(1, 0, 123)
    worksheet.write(1, 1, 123.456)

    # Insert an image.
    #worksheet.insert_image('B5', 'logo.png')

    workbook.close()

def demo_reader():
    # Give the location of the file
    path = "C:\\Users\\Anastasia Siedykh\\Documents\\Backup\\KPI report\\MODULE SET V6\\0.new_629_report_2021.xlsx"

    # To open the workbook
    # workbook object is created
    wb_obj = openpyxl.load_workbook(path)

    # Get workbook active sheet object
    # from the active attribute
    sheet_obj = wb_obj.active


    # Cell objects also have a row, column,
    # and coordinate attributes that provide
    # location information for the cell.

    # Note: The first row or
    # column integer is 1, not 0.

    # Cell object is created by using
    # sheet object's cell() method.
    rows_count = str(sheet_obj.calculate_dimension()).rsplit(':')

    rows_count = int(str(rows_count[1])[2:])
    print(rows_count)
    string = []
    classified_base_2021 = []
    for row in range(1, rows_count+1):
        str_ = []
        for col in range(1,33):
            cell_obj = sheet_obj.cell(row=row, column=col)
            str_.append(cell_obj.value)
        string.append(str_)


    for i in string:
        x = CBase_2021_quadra_workout()
        string_class = x.classify_base_from_xlxs(i)
        classified_base_2021.append(string_class)
    total = 0
    for z in classified_base_2021[1:]:
        for d in z:
            total += float(d.sales_euro)
    print(total)







    # Print value of cell object
    # using the value attribute


if __name__ == '__main__':
    demo_reader()